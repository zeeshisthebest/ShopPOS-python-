import datetime
from tkinter.messagebox import showinfo

import win32api
import win32print
from openpyxl import Workbook as wB
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image

# All the styles
main_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,
                           wrap_text=True, shrink_to_fit=False, indent=0)
title_alignment = Alignment(horizontal='right', vertical='center', text_rotation=0,
                            wrap_text=True, shrink_to_fit=False, indent=0)
date_alignment = Alignment(horizontal='right', vertical='center', text_rotation=0,
                           wrap_text=True, shrink_to_fit=False, indent=0)
price_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,
                            wrap_text=False, shrink_to_fit=False, indent=0)
customer_alignment = Alignment(horizontal='left', vertical='center', text_rotation=0,
                               wrap_text=False, shrink_to_fit=False, indent=0)

title_fill_color = '00CC00FF'
company_name_fill = PatternFill(start_color=title_fill_color, end_color=title_fill_color,
                                fill_type='solid')
side = Side(style='thin')
thin_border = Border(left=side, right=side, top=side, bottom=side)
total_border = Border(left=None, right=None,
                      top=Side(style='double'), bottom=None)

arial_font_name = 'Verdana'
head_font = Font(size=10, name=arial_font_name)
arial_font = Font(size=10, name=arial_font_name)
footer_font_color = '666666'


# Styles finished /////////////////////////////////////////////////////////////

class GenerateReceipt:
    def __init__(self, invoice, path):
        self.__invoice = invoice
        self.__title = f"Invoice {invoice} {datetime.date.today()}"
        self.xls = wB()
        self.__sheet = self.xls.active
        self.filename = self.__title + ".xlsx"
        self.path = path
        self.__row_num = 10
        # adjusting the widths of the Columns
        self.__sheet.column_dimensions['A'].width = 5
        self.__sheet.column_dimensions['B'].width = 12
        self.__sheet.row_dimensions[1].height = 25

        # Setting the Header
        header_odd = self.__sheet.oddHeader.center
        header_odd.text = 'RECEIPT'
        header_odd.size = 12
        header_odd.font = arial_font_name
        header_odd.color = '000000'

        header_even = self.__sheet.evenHeader.center
        header_even.text = 'RECEIPT'
        header_even.size = 12
        header_even.font = arial_font_name
        header_even.color = '000000'

        header_oddInv = self.__sheet.oddHeader.left
        header_oddInv.text = f'Invoice: {self.__invoice}'
        header_oddInv.size = 10
        header_oddInv.font = arial_font_name
        header_oddInv.color = '000000'

        header_evenInv = self.__sheet.evenHeader.left
        header_evenInv.text = f'Invoice: {self.__invoice}'
        header_evenInv.size = 10
        header_evenInv.font = arial_font_name
        header_evenInv.color = '000000'

        # Setting the Footer
        footer_right_text = 'Page &[Page] of &N.'
        fb_link = "www.fb.com/SaleemChemicals"
        footer_center_text = "Receipt required for return and exchange.\n" \
                             "Terms and Conditions applied."

        # Left Footer
        self.__sheet.oddFooter.left.text = fb_link
        self.__sheet.oddFooter.left.color = footer_font_color
        self.__sheet.oddFooter.left.font = arial_font_name
        self.__sheet.oddFooter.left.size = 8
        self.__sheet.evenFooter.left.text = fb_link
        self.__sheet.evenFooter.left.color = footer_font_color
        self.__sheet.evenFooter.left.font = arial_font_name
        self.__sheet.evenFooter.left.size = 8

        # Right Footer
        footer_oddR = self.__sheet.oddFooter.right
        footer_oddR.text = footer_right_text
        footer_oddR.size = 8
        footer_oddR.font = arial_font_name
        footer_oddR.color = footer_font_color

        footer_evenR = self.__sheet.evenFooter.right
        footer_evenR.text = footer_right_text
        footer_evenR.size = 8
        footer_evenR.font = arial_font_name
        footer_evenR.color = footer_font_color

        # Center Footer
        footer_oddC = self.__sheet.oddFooter.center
        footer_oddC.text = footer_center_text
        footer_oddC.size = 9
        footer_oddC.font = arial_font_name
        footer_oddC.color = footer_font_color

        footer_evenC = self.__sheet.evenFooter.center
        footer_evenC.text = footer_center_text
        footer_evenC.size = 9
        footer_evenC.font = arial_font_name
        footer_evenC.color = footer_font_color
        # Footers finished ////////////////////////////////////////////////////

        # Head Info
        self.__sheet.merge_cells('A1:J3')
        title_cell = self.__sheet['A1']
        # title_cell.fill = company_name_fill
        title_cell.value = "0333 - 2384042\n0332 - 2569183\nOffice# 17, B - Road\nLiaquatabad, Karachi"
        img = Image('logo.png')
        img.width = 663 / 7
        img.height = 392 / 7
        self.__sheet.add_image(img, 'A1')
        title_cell.alignment = title_alignment
        title_cell.font = head_font

        # Date of the day
        date_cell_range = 5
        self.__sheet.merge_cells(f'H{date_cell_range}:J{date_cell_range}')
        self.__sheet[f'H{date_cell_range}'] = "Date: " + str(datetime.date.today())
        self.__sheet[f'H{date_cell_range}'].alignment = date_alignment
        self.__sheet[f'H{date_cell_range}'].font = head_font

        # Items list header
        self.__sheet['A9'] = '#'
        self.__sheet['B9'] = 'Items'
        self.__sheet['H9'] = 'Unit'
        self.__sheet['I9'] = 'Quantity'
        self.__sheet['J9'] = 'Total'

        for a in "ABHIJ":
            self.__sheet[f'{a}9'].alignment = main_alignment
            self.__sheet[f'{a}9'].border = thin_border
            self.__sheet[f'{a}9'].font = arial_font

        self.__sheet.merge_cells('B9:G9')
        # Item list head finished//////////////////////////////////////////

    def set_customer_name(self, customer, cell, address):
        customer_cell = self.__sheet['A5']
        customer_cell.value = "Customer: " + customer.title()

        contact_cell = self.__sheet['A6']
        contact_cell.value = "Contact: " + (cell if cell != "" else "----")

        address_cell = self.__sheet['A7']
        address_cell.value = "Address: " + address.title()

        customer_cell.alignment = customer_alignment
        customer_cell.font = arial_font

        contact_cell.alignment = customer_alignment
        contact_cell.font = arial_font

        address_cell.alignment = customer_alignment
        address_cell.font = arial_font

        self.__sheet.merge_cells('A5:G5')
        self.__sheet.merge_cells('A6:G6')
        self.__sheet.merge_cells('A7:G7')

    def print_list(self, item_list):
        offset = 9
        grand_total = 0
        # Inserting the list data
        for key in range(0, len(item_list) - 5):
            row = item_list[key]
            str_row_num = str(self.__row_num)
            merging_row = 'B' + str_row_num + ':' + 'G' + str_row_num
            self.__sheet.merge_cells(merging_row)

            # Creating cell ids
            serial_number = 'A' + str_row_num
            item_cell_number = 'B' + str_row_num
            unit_cell_number = 'H' + str_row_num
            quantity_cell_number = 'I' + str_row_num
            total_cell_number = 'J' + str_row_num

            # Writing the data in the cells
            # serial number
            # self.__sheet[serial_number].value = self.__row_num - offset ------------------------------
            self.__sheet[serial_number].value = row['serial']
            self.__sheet[serial_number].alignment = main_alignment
            self.__sheet[serial_number].border = thin_border
            self.__sheet[serial_number].font = arial_font
            #  Items cell
            # self.__sheet[item_cell_number].value = '    ' + row[0].title() ---------------------------
            self.__sheet[item_cell_number].value = f"    {row['item'].title()} ({row['measure']})"
            self.__sheet[item_cell_number].border = thin_border
            self.__sheet[item_cell_number].font = arial_font
            self.__sheet[f'C{self.__row_num}'].border = thin_border
            self.__sheet[f'D{self.__row_num}'].border = thin_border
            self.__sheet[f'E{self.__row_num}'].border = thin_border
            self.__sheet[f'F{self.__row_num}'].border = thin_border
            self.__sheet[f'G{self.__row_num}'].border = thin_border

            # unit cell
            # self.__sheet[unit_cell_number].value = row[1] ----------------------
            self.__sheet[unit_cell_number].value = row['unit']
            self.__sheet[unit_cell_number].border = thin_border
            self.__sheet[unit_cell_number].font = arial_font
            self.__sheet[total_cell_number].alignment = date_alignment

            # quantity cell
            # self.__sheet[quantity_cell_number].value = row[2] ----------------------------
            self.__sheet[quantity_cell_number].value = row['quantity']
            self.__sheet[quantity_cell_number].border = thin_border
            self.__sheet[quantity_cell_number].font = arial_font
            self.__sheet[total_cell_number].alignment = date_alignment

            # total cell
            # grand_total += row[2] * row[1]

            # self.__sheet[total_cell_number].value = row[2] * row[1] ------------------------
            self.__sheet[total_cell_number].value = row['total']
            self.__sheet[total_cell_number].border = thin_border
            self.__sheet[total_cell_number].font = arial_font
            self.__sheet[total_cell_number].alignment = date_alignment

            # To iterate through the rows
            self.__row_num += 1

        # self.print_the_total(grand_total)

    def print_the_total(self, total_grand):
        self.__row_num += 1
        grand_total_cells_to_merge = f'A{self.__row_num}:H{self.__row_num}'
        self.__sheet.merge_cells(grand_total_cells_to_merge)
        self.__sheet.merge_cells(f'I{self.__row_num}:J{self.__row_num}')
        self.__sheet[f'A{self.__row_num}'].value = "TOTAL: "
        self.__sheet[f'A{self.__row_num}'].font = arial_font
        self.__sheet[f'I{self.__row_num}'].value = total_grand
        self.__sheet[f'I{self.__row_num}'].font = arial_font
        self.__sheet[f'I{self.__row_num}'].alignment = date_alignment

        for alpha in "ABCDEFGHIJ":
            self.__sheet[f'{alpha}{self.__row_num}'].border = total_border

        self.__sheet[f'A{self.__row_num}'].alignment = date_alignment
        self.__sheet[f'J{self.__row_num}'].alignment = price_alignment

    def save_excel(self):

        print_size = f'A1:J{self.__row_num}'
        self.__sheet.print_options.horizontalCentered = True
        self.__sheet.print_options.verticalCentered = False
        self.__sheet.print_area = print_size
        self.__sheet.page_setup.paperSize = self.__sheet.PAPERSIZE_A4
        self.__sheet.sheet_properties.pageSetUpPr.fitToPage = True
        self.__sheet.sheet_properties.pageSetUpPr.autoPageBreaks = True
        self.__sheet.page_setup.fitToWidth = True
        self.__sheet.page_setup.fitToHeight = False
        self.__sheet.page_margins.left = 0.5
        self.__sheet.page_margins.right = 0.5
        self.__sheet.page_margins.top = 1.3
        self.__sheet.page_margins.header = 0.5
        self.__sheet.page_margins.bottom = 1

        try:
            print(self.path)
            self.xls.save(self.path)

        except Exception:
            print("saving error", Exception)
            return False
        return True

    def print_to_pdf(self):
        try:
            win32api.ShellExecute(0, "printto", self.path, '"%s"' % win32print.GetDefaultPrinter(),
                                  ".", 0)
        except Exception:
            showinfo("Error", "There was error in printing as pdf")
